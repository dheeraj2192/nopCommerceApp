import openpyxl

def rowCount(file,sheetname):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetname]
    return sheet.max_row

def colCount(file,sheetname):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetname]
    return sheet.max_column


def readExcel(file,sheetname,r,c):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetname]
    return sheet.cell(r,c).value

def writeExcel(file,sheetname,r,c,data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetname]
    sheet.cell(r,c).value=data
    wb.save(file)